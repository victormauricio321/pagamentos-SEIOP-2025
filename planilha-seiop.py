import streamlit as st
import pandas as pd

from io import BytesIO
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------
# TÍTULO DO APP
# ----------------------------------------------------------
st.title("Arquivo CSV → Excel (.xlsx)")

# Passo 1: Upload do arquivo CSV
uploaded_file = st.file_uploader("Arraste ou selecione um arquivo CSV", type=["csv"])

if uploaded_file is not None:
    # Leitura do CSV
    df = pd.read_csv(uploaded_file, encoding='latin1')
    
    #coluna_esquerda, coluna_direita = st.columns(2)

    #with coluna_esquerda:
        #st.subheader("Pré-visualização dos dados originais")
        #st.dataframe(df.head())

    # ----------------------------------------------------------
    # Passo 2: EXEMPLO DE TRATAMENTO DE DADOS
    # ----------------------------------------------------------
    df = df.drop(["Unidade Orçamentária",
               "Unidade Gestora",
               "Fonte",
               "Ação",
               "Nota de Empenho",
               "Nota de Liquidação",
               "Desp. Empenhadas",
               "Desp. Emp. a Liquidar",
               "Empenhos Cancelados",
               "Desp. Emp. em Liquidação",
               "Desp. Liquidadas",
               "Despesas Liquidadas a Pagar",
                ],
                axis=1
             )
    df[['Número Automático', 'Objeto']] = df['Contrato'].str.extract(r'(.{11})(.*)')
    df.drop(columns='Contrato', inplace=True)

    nova_ordem = ['Número Automático', 'Objeto'] + [col for col in df.columns if col not in ['Número Automático', 'Objeto']]
    df = df[nova_ordem]

    df = df.drop(["Objeto"], axis=1)
    df = df.drop(["Processo"], axis=1)

    df["Número Automático"] = df["Número Automático"].str[:8]

    df[['CNPJ', 'Nome Credor']] = df['Credor'].str.extract(r'(.{17})(.*)')
    df.drop(columns='Credor', inplace=True)

    nova_ordem_colunas = ['Nome Credor',
                 'Número Automático',
                 'Número Original',
                 'Subelemento',
                 'Ordem Bancária',
                 'Desp. Pagas',
                 'RP Pagos'
                ]

    df = df[nova_ordem_colunas]

    df = df[df['Nome Credor'] != 'Crater construções ltda'].reset_index(drop=True)

    df = df.sort_values(by=['Nome Credor', 'Número Automático'], ascending=[True, True])

    df['Subelemento'] = pd.to_numeric(df['Subelemento'], errors='coerce').astype('Int64')
    df['Número Automático'] = pd.to_numeric(df['Número Automático'], errors='coerce').astype('Int64')

    df = df.loc[df['Subelemento'] > 40000000]
    df = df.reset_index(drop=True)

    #with coluna_direita:
    st.subheader("Pré-visualização dos dados tratados")
    st.dataframe(df)

    # ----------------------------------------------------------
    # Passo 3: EXPORTAR PARA XLSX
    # ----------------------------------------------------------
    def converter_para_excel(dataframe):
        from openpyxl.utils import get_column_letter

        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            dataframe.to_excel(writer, index=False, sheet_name="Tratado")

            # Ajusta largura das colunas automaticamente
            worksheet = writer.sheets["Tratado"]
            for i, col in enumerate(dataframe.columns):
                # Calcula a largura com base no tamanho máximo entre nome e conteúdo
                max_len = max(
                    dataframe[col].astype(str).map(len).max(),
                    len(col)
                ) + 2  # margem
                worksheet.column_dimensions[get_column_letter(i + 1)].width = max_len

        return buffer.getvalue()

    excel_bytes = converter_para_excel(df)

    # Botão para download
    st.download_button(
        label="⬇️  Baixar arquivo em Excel",
        data=excel_bytes,
        file_name="pagamentos-SEIOP-2025.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )